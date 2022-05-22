import enum
import json
import openpyxl as exl
from openpyxl.styles import Font
from math import sqrt, floor


METHODS = {
    "FR": "F-R", "PH": "PH", "graphtpp": "graphTPP", "drgraph": "DRGraph",
    "graphtsne": "GraphTSNE", # "EGL(node2vec)": "GEGraph(node2vec)", "EGL": "GEGraph(node2vec-a)",
    "EGL(node2vec,class)": "GEGraph(node2vec)", "EGL(class)": "GEGraph(node2vec-a)",
}
METRICS = {
    "node_spread": "Node Spread", "node_occlusions": "Node Occlusions", 
    "edge_crossings": "Edge Crosssings", "edge_crossings_outside": "Edge Crossings(Outside)",
    "minimum_angle": "Minimum Angle", "edge_length_variation": "Edge Length Variation", 
    "group_overlap": "Group Overlap", "entropy": "Entropy", "autocorrelation": "Autocorrelation"
}
GRAPHS = {
    "miserables": "Les Mis{\'e}rables",
    "cornell": "Webkb", 
    "facebook": "Facebook", 
    "science": "Science", 
    "cora": "Cora", 
    "citeseer": "Citeseer"
}
# GRAPHS = ["miserables", "facebook", "science"]

graph_results = {}
for graph in GRAPHS.keys():
    json_path = "./{}_layout_evaluation.json".format(graph)
    with open(json_path, "r") as f:
        text = f.read()
        res = json.loads(text)
        graph_results[graph] = res

for graph, result in graph_results.items():
    book = exl.Workbook()
    sheet = book['Sheet']
    header = ["Graph", "Method"]
    for m in METRICS.values():
        header.append(m)
    for i in range(len(header)):
        sheet.cell(1, i + 1).value = header[i]
    row = 2
    for graph in GRAPHS.keys():
        sheet.cell(row, 1).value = "\multirow{7}*{" + GRAPHS[graph] + "}"
        for method in METHODS.keys():
            # sheet.cell(row, 1).value = "ababab"
            sheet.cell(row, 2).value = METHODS[method]
            if method in graph_results[graph].keys():
                for i, metric in enumerate(METRICS.keys()):
                    if metric in graph_results[graph][method].keys():
                        if graph_results[graph][method][metric] != -1:
                            sheet.cell(row, 3 + i).value = "{0:.3f}".format(graph_results[graph][method][metric])
                        else:
                            sheet.cell(row, 3 + i).value = "-"
                    else:
                        sheet.cell(row, 3 + i).value = "-"
                        sheet.cell(row, 3 + i).font = Font(bold=True)
            else:
                for i, metric in enumerate(METRICS.keys()):
                    sheet.cell(row, 3 + i).value = "-"
                    continue

            row += 1
    for r in [2, 9, 16, 23, 30, 37]:
        for c in ["C", "D", "E", "F", "G", "H", "E", "F", "G", "H", "I", "J", "K"]:
            min_val = 100
            total = 0
            count = 0
            vals = []
            for i in range(r, r + 7):
                val = sheet[c + str(i)].value
                if val == "-" or val == "" or val is None:
                    continue
                num_val = float(val)
                vals.append(num_val)
                if num_val < min_val:
                    min_val = num_val
            if vals:
                vals.sort()
                # q1 = sqrt(vals[1])
                # q3 = sqrt(vals[-2])
                q1 = vals[floor(len(vals) / 4)]
                if q1 >= 1:
                    q1 = sqrt(q1)
                q3 = vals[floor(len(vals) * 3 / 4)]
                if q3 >= 1:
                    q3 = sqrt(q3)
                iqr = q3 - q1
                ub = q3 + iqr * 1.5
                # print(vals, q1, q3, iqr, ub)
            else:
                ub = None
            for i in range(r, r + 7):
                val = sheet[c + str(i)].value
                if val == "-" or val == "" or val is None:
                    continue
                num_val = float(val)
                if num_val == min_val:
                    sheet[c + str(i)].font = Font(bold=True)
                # elif num_val == vals[1]:
                #     # print("here")
                #     sheet[c + str(i)].font = Font(italic=True)
                if ub and num_val >= ub:
                    sheet[c + str(i)].font = Font(bold=True, color="FF0000")

    book.save("evaluation_result4.xlsx")